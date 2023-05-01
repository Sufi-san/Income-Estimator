from tkinter import *
from tkinter import filedialog, messagebox, ttk
import tkinter.font as tkfont
from PIL import ImageTk, Image
import webbrowser
import os
import openpyxl
from openpyxl.styles import Alignment, Font


def close_win():
    if window_counter1 > 0:
        if messagebox.askokcancel("Close All Windows?", "All unsaved data will be lost, do you want to continue?"):
            root.destroy()
    else:
        root.destroy()


def open_chatbot():
    webbrowser.open_new_tab("Chatbot.html")


def browse_folder():
    filepath = filedialog.askopenfilename(initialdir="Saved Files", title="Select a File", filetypes=(("All files", "*.*"), ("all files", "*.*")))
    os.startfile(filepath)


def open_create():

    def set_reset():
        total_collection = entry_1.get()
        if total_collection:
            if lbl_3.cget('text') != '-':
                col_excluding_prof = float(total_collection) - float(lbl_3.cget('text'))
            else:
                col_excluding_prof = float(total_collection)
            lbl_5.configure(text=col_excluding_prof)
        else:
            messagebox.showinfo('Details Incomplete', 'Please Enter the total collection.')

    def estimate_inc():
        global emp_prod_profit
        col_excluding_prof = 0
        if entry_1.get() != '':
            col_excluding_prof = round_to_3(float(lbl_5.cget('text')))
        if tree_view3.get_children() and col_excluding_prof != 0:
            lbl_20.configure(fg="#313232")
            for index in tree_view3.get_children():
                value_lst = tree_view3.item(index)['values']
                if value_lst[3] == 0:
                    if str(value_lst[2]).find('%') != -1:
                        percent_salary = float(col_excluding_prof * float(value_lst[2][:str(value_lst[2]).find('%')])/100)
                        tree_view3.set(index, 'Estimated Income', round_to_3(percent_salary))
                    else:
                        tree_view3.set(index, 'Estimated Income', value_lst[2])
                else:
                    est_inc = 0
                    for item in emp_prod_profit:
                        if str(value_lst[0]) == str(item[0]):
                            if str(value_lst[2]).find('%') != -1:
                                percent_salary = float(
                                    col_excluding_prof * float(value_lst[2][:str(value_lst[2]).find('%')]) / 100)
                                est_inc = round_to_3(percent_salary + float(item[1]))
                                break
                            else:
                                est_inc = round_to_3(float(value_lst[2]) + float(item[1]))
                                break
                    tree_view3.set(index, 'Estimated Income', est_inc)
            amt_distributed = 0
            for index in tree_view3.get_children():
                new_value_lst = tree_view3.item(index)['values']
                amt_distributed += float(new_value_lst[4])
            amt_after_distribution = float(entry_1.get()) - float(amt_distributed)
            lbl_16.configure(text=round_to_3(amt_after_distribution))

        else:
            messagebox.showinfo('Detatils Incomplete', 'Please fill all required fields.')

    def save_file():
        filename = str(entry_10.get())
        if filename == '':
            messagebox.showinfo('No Name', "Please Enter a name for the file.")
        elif os.path.isfile("Saved Files/"+filename+".xlsx"):
            messagebox.showinfo('File Exists', "File with same name already exists, please enter a different name.")
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            column_names = ["ID", "Name", "Salary", "Products Sold", "Estimated Income"]
            ws.append(column_names)
            for index in tree_view3.get_children():
                item = tree_view3.item(index)['values']
                ws.append(item)
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 29
            ws.column_dimensions['G'].width = 28
            ws.column_dimensions['H'].width = 18
            i = 0
            for i in range(len(tree_view3.get_children())+1):
                i += 1
                ws.cell(row=i, column=1).alignment = Alignment(horizontal='left')
                ws.cell(row=i, column=2).alignment = Alignment(horizontal='left')
                ws.cell(row=i, column=3).alignment = Alignment(horizontal='center')
                ws.cell(row=i, column=4).alignment = Alignment(horizontal='center')
                ws.cell(row=i, column=5).alignment = Alignment(horizontal='center')
            ws.cell(row=1, column=7, value="Total Collection:")
            ws.cell(row=2, column=7, value="Profit from Product Sale:")
            ws.cell(row=3, column=7, value="Collection excluding Products:")
            ws.cell(row=4, column=7, value="Amount Left after Distribution:")
            ws.cell(row=1, column=8, value=str(entry_1.get()))
            ws.cell(row=2, column=8, value=lbl_3.cget('text'))
            ws.cell(row=3, column=8, value=lbl_5.cget('text'))
            ws.cell(row=4, column=8, value=lbl_16.cget('text'))
            bold_font = Font(bold=True)
            for i in range(5):
                i += 1
                ws.cell(row=i, column=7).font = bold_font
                ws.cell(row=i, column=8).font = bold_font
                ws.cell(row=i, column=8).alignment = Alignment(horizontal='center')
            for cell in ws[1]:
                cell.font = bold_font
            wb.save("Saved Files/"+filename+".xlsx")
            messagebox.showinfo("File Saved", "The file is successfully saved in 'Saved Files' folder.")

    def unique_check(value):
        if value == 1:
            exists1 = False
            for record in tree_view1.get_children():
                if str(cmb_box1.get()) == str(tree_view1.item(record)['values'][0]):
                    exists1 = True
                    break
            return exists1
        elif value == 2:
            exists2 = False
            for record in tree_view2.get_children():
                if str(entry_6.get()) == str(tree_view2.item(record)['values'][0]):
                    exists2 = True
                    break
            return exists2
        elif value == 3:
            exists3 = False
            for record in tree_view3.get_children():
                if str(entry_2.get()) == str(tree_view3.item(record)['values'][0]):
                    exists3 = True
                    break
            return exists3

    def close_action():
        messagebox_toplevel = Toplevel(root)
        messagebox_toplevel.title("Invisible")
        messagebox_toplevel.geometry("0x0+0+0")
        messagebox_toplevel.resizable(False, False)
        messagebox_toplevel.iconbitmap("Required Images/app_icon.ico")
        messagebox_toplevel.overrideredirect(True)
        messagebox_toplevel.grab_set()
        if messagebox.askokcancel("Close Window?", "All unsaved data will be lost, do you want to continue?"):
            global window_counter1, count_tree1, count_tree2, count_tree3, products, emp_prod_profit
            count_tree1 = count_tree2 = count_tree3 = 0
            products = emp_prod_profit = []
            window_counter1 -= 1
            child.destroy()
            root.deiconify()
        messagebox_toplevel.grab_release()
        messagebox_toplevel.destroy()

    def when_focused1(event):
        check = entry_4.get()
        if check == placeholder_ent4a or check == placeholder_ent4b or check == "":
            event.widget.delete(0, END)
            entry_4.configure(fg="black")
            if v2.get() == 1:
                entry_4.configure(validatecommand=(validate1, "%P"))
            elif v2.get() == 2:
                entry_4.configure(validatecommand=(validate2, "%P"))

    def enable_prod():
        if v1.get() == 1:
            entry_5.configure(state=NORMAL)
            entry_6.configure(state=NORMAL)
            entry_7.configure(state=NORMAL)
            entry_8.configure(state=NORMAL)
            entry_9.configure(state=NORMAL)
            btn_7.configure(state=NORMAL)
            btn_8.configure(state=NORMAL)
            btn_9.configure(state=NORMAL)
            btn_10.configure(state=NORMAL)
            btn_11.configure(state=NORMAL)
            btn_12.configure(state=NORMAL)
            btn_13.configure(state=NORMAL)
            btn_14.configure(state=NORMAL)
            cmb_box1.configure(state="readonly")
            lbl_2.configure(state=NORMAL)
            lbl_3.configure(state=NORMAL)
            lbl_9.configure(state=NORMAL)
            lbl_10.configure(state=NORMAL)
            lbl_11.configure(state=NORMAL)
            lbl_12.configure(state=NORMAL)
            lbl_13.configure(state=NORMAL)
            lbl_14.configure(state=NORMAL)
        elif v1.get() == 0:
            entry_5.delete(0, END)
            entry_6.delete(0, END)
            entry_7.delete(0, END)
            entry_8.delete(0, END)
            entry_9.delete(0, END)
            entry_5.configure(state=DISABLED)
            entry_6.configure(state=DISABLED)
            entry_7.configure(state=DISABLED)
            entry_8.configure(state=DISABLED)
            entry_9.configure(state=DISABLED)
            btn_7.configure(state=DISABLED)
            btn_8.configure(state=DISABLED)
            btn_9.configure(state=DISABLED)
            btn_10.configure(state=DISABLED)
            btn_11.configure(state=DISABLED)
            btn_12.configure(state=DISABLED)
            btn_13.configure(state=DISABLED)
            btn_14.configure(state=DISABLED)
            cmb_box1.configure(state=DISABLED)
            cmb_box1.set("")
            lbl_2.configure(state=DISABLED)
            lbl_3.configure(state=DISABLED)
            lbl_9.configure(state=DISABLED)
            lbl_10.configure(state=DISABLED)
            lbl_11.configure(state=DISABLED)
            lbl_12.configure(state=DISABLED)
            lbl_13.configure(state=DISABLED)
            lbl_14.configure(state=DISABLED)
            rem_all_rec_tree(1)

    def on_combo_select(event):
        global products
        font = tkfont.nametofont(str(event.widget.cget('font')))
        max_num = 0
        if products:
            for product in products:
                if len(product) > max_num:
                    max_num = products.index(product)
            width = font.measure(products[max_num])
            if width > cmb_box1.winfo_width():
                style = ttk.Style()
                style.configure('TCombobox', postoffset=(0, 0, width, 0))

    def choose():
        entry_3.focus_set()
        if v2.get() == 1:
            entry_4.delete(0, END)
            entry_4.configure(state=NORMAL)
            if entry_4.get() == "":
                entry_4.configure(fg="#b4b4b4")
                entry_4.insert(0, placeholder_ent4a)
        elif v2.get() == 2:
            entry_4.delete(0, END)
            entry_4.configure(state=NORMAL)
            if entry_4.get() == "":
                entry_4.configure(fg="#b4b4b4")
                entry_4.insert(0, placeholder_ent4b)

    def fill_data_msg(num, title, message):
        messagebox_toplevel = Toplevel(root)
        messagebox_toplevel.title("Invisible")
        messagebox_toplevel.geometry("0x0+0+0")
        messagebox_toplevel.resizable(False, False)
        messagebox_toplevel.iconbitmap("Required Images/app_icon.ico")
        messagebox_toplevel.overrideredirect(True)
        messagebox_toplevel.grab_set()
        if num == 1:
            messagebox.showinfo(title, message)
        elif num == 2:
            messagebox.showerror(title, message)
        messagebox_toplevel.grab_release()
        messagebox_toplevel.destroy()

    def add_rec_tree(a):
        global count_tree1, count_tree2, count_tree3
        global products, emp_prod_profit
        if a == 1:
            if cmb_box1.get() != "" and entry_5.get() != "":
                if unique_check(1):
                    messagebox.showinfo('Identical Values', 'Please Select unique Product.')
                else:
                    t_profit = 0
                    for index in tree_view2.get_children():
                        tree2_list = tree_view2.item(index)["values"]
                        if cmb_box1.get() == tree2_list[0]:
                            t_profit = round_to_3(float(tree2_list[2]) * int(entry_5.get()))
                    value = (cmb_box1.get(), int(entry_5.get()), t_profit)
                    tag = ('bluerow',)
                    tree_view1.insert(parent="", index="end", iid=str(count_tree1), text="", values=value, tags=tag)
                    count_tree1 += 1
            else:
                fill_data_msg(1, "Data Incomplete", "Please fill all the required details.")
        if a == 2:
            if entry_6.get() != "" and entry_7.get() != "" and entry_8.get() != "" and entry_9.get() != "":
                if unique_check(2):
                    messagebox.showinfo('Identical Values', 'Please Enter Unique Product Name.')
                else:
                    margin = entry_7.get() + "% of Profit = " + str(round_to_3(float(entry_7.get()) * (float(entry_9.get()) - float(entry_8.get())) / 100))
                    profit = round_to_3(float(entry_9.get()) - float(entry_8.get()))
                    value = (entry_6.get(), margin, profit)
                    products.append(entry_6.get())
                    cmb_box1.config(values=products)
                    tag = ('bluerow',)
                    tree_view2.insert(parent="", index="end", iid=str(count_tree2), text="", values=value, tags=tag)
                    count_tree2 += 1
            else:
                fill_data_msg(1, "Data Incomplete", "Please fill all the required details.")
        if a == 3:
            if entry_2.get() != "" and entry_3.get() != "" and entry_4.get() != "":
                if unique_check(3):
                    messagebox.showinfo('Identical Values', 'Please Enter Unique Employee ID.')
                else:
                    lbl_20.configure(fg="yellow")
                    data = 0
                    tot_prod = 0
                    if v2.get() == 1:
                        data = entry_4.get()
                    elif v2.get() == 2:
                        data = str(entry_4.get()) + "% of Collection"
                    if v1.get() == 1 and tree_view1.get_children() != []:
                        for record in tree_view1.get_children():
                            tot_prod = tot_prod + int(tree_view1.item(record)["values"][1])
                    value = (str(entry_2.get()), str(entry_3.get()), data, tot_prod, "-")
                    tag = ('bluerow',)
                    tree_view3.insert(parent="", index="end", iid=str(count_tree3), text="", values=value, tags=tag)
                    count_tree3 += 1
                    if tot_prod != 0:
                        chk_btn1.configure(state=DISABLED)
                        emp_prod_profit.append([str(entry_2.get())])
                        emp_profit = 0
                        tot_prod_profit = 0
                        for index1 in tree_view1.get_children():
                            for index2 in tree_view2.get_children():
                                prod_real = tree_view1.item(index1)['values']
                                prod_img = tree_view2.item(index2)['values']
                                if prod_real[0] == prod_img[0]:
                                    emp_profit = emp_profit + float(prod_real[2])*float(prod_img[1][:prod_img[1].index('%')])/100
                                    break
                        for item in emp_prod_profit:
                            if item[0] == str(entry_2.get()):
                                emp_prod_profit[emp_prod_profit.index(item)].append(emp_profit)
                        for index1 in tree_view1.get_children():
                            tot_prod_profit += float(tree_view1.item(index1)['values'][2])
                        round_to_3(tot_prod_profit)
                        for item in emp_prod_profit:
                            if item[0] == str(entry_2.get()):
                                emp_prod_profit[emp_prod_profit.index(item)].append(tot_prod_profit)
                                break
                        final_prod_profit = 0
                        for i in range(len(emp_prod_profit)):
                            final_prod_profit += emp_prod_profit[i][2]
                            i += 1
                        lbl_3.configure(text=final_prod_profit)
                        if entry_1.get() != '':
                            col_excluding_prof = float(entry_1.get()) - float(lbl_3.cget('text'))
                            lbl_5.configure(text=col_excluding_prof)
            else:
                fill_data_msg(1, "Data Incomplete", "Please fill all the required details.")

    def upd_rec_tree(a):
        global products, emp_prod_profit
        if a == 1:
            record_tuple = tree_view1.selection()
            record = record_tuple[0]
            if cmb_box1.get() != "" and entry_5.get() != "":
                if unique_check(1) and str(tree_view1.item(record)['values'][0]) != str(cmb_box1.get()):
                    messagebox.showinfo('Identical Values', 'Please Select unique Product.')
                else:
                    t_profit = 0
                    for index in tree_view2.get_children():
                        tree2_list = tree_view2.item(index)["values"]
                        if cmb_box1.get() == tree2_list[0]:
                            t_profit = float(tree2_list[2]) * int(entry_5.get())
                    value = (cmb_box1.get(), int(entry_5.get()), t_profit)
                    tree_view1.item(record, values=value)
            else:
                fill_data_msg(1, "Data Incomplete", "Please fill all the required details.")
        if a == 2:
            record_tuple = tree_view2.selection()
            record = record_tuple[0]
            old_value = tree_view2.item(record)["values"][0]
            upd_idx = products.index(old_value)
            if entry_6.get() != "" and entry_7.get() != "" and entry_8.get() != "" and entry_9.get() != "":
                if unique_check(2) and str(tree_view2.item(record)['values'][0]) != str(entry_6.get()):
                    messagebox.showinfo('Identical Values', 'Please Enter Unique Product Name.')
                else:
                    margin = entry_7.get() + "% of Profit = " + str(round_to_3(float(entry_7.get())*(float(entry_9.get()) - float(entry_8.get()))/100))
                    profit = round_to_3(float(entry_9.get()) - float(entry_8.get()))
                    value = (entry_6.get(), margin, profit)
                    tree_view2.item(record, values=value)
                    for child_tr2 in tree_view1.get_children():
                        if tree_view1.item(child_tr2)['values'][0] == old_value:
                            tree_view1.set(child_tr2, 'Product', entry_6.get())
                            new_tot_prof = float(tree_view1.item(child_tr2)['values'][1])*float(tree_view2.item(record)['values'][2])
                            tree_view1.set(child_tr2, 'Total Profit', new_tot_prof)
                    products[upd_idx] = entry_6.get()
                    cmb_box1.set('')
                    cmb_box1.config(values=products)
            else:
                fill_data_msg(1, "Data Incomplete", "Please fill all the required details.")
        if a == 3:
            record_tuple = tree_view3.selection()
            record = record_tuple[0]
            if entry_2.get() != "" and entry_3.get() != "" and entry_4.get() != "":
                if unique_check(3) and str(tree_view3.item(record)['values'][0]) != str(entry_2.get()):
                    messagebox.showinfo('Identical Values', 'Please Enter Unique Employee ID.')
                else:
                    lbl_20.configure(fg="yellow")
                    data = 0
                    tot_prod = 0
                    no_prods = True
                    if v2.get() == 1:
                        data = entry_4.get()
                    elif v2.get() == 2:
                        data = str(entry_4.get()) + "% of Collection"
                    if v1.get() == 1 and tree_view1.get_children() != []:
                        for index in tree_view1.get_children():
                            tot_prod = tot_prod + int(tree_view1.item(index)["values"][1])
                    value = (str(entry_2.get()), str(entry_3.get()), data, tot_prod, "-")
                    tree_view3.item(record, values=value)
                    print("EMP_PROD_PROFIT: ", emp_prod_profit)
                    for item in emp_prod_profit:
                        if item[0] == str(entry_2.get()):
                            print(item[0], " = ", str(entry_2.get()))
                            emp_prod_profit.remove(emp_prod_profit[emp_prod_profit.index(item)])
                    print("EMP_PROD_PROFIT: ", emp_prod_profit)
                    for index in tree_view3.get_children():
                        if tree_view3.item(index)['values'][3] != 0:
                            no_prods = False
                            break
                    if no_prods:
                        chk_btn1.configure(state=NORMAL)
                    if tot_prod != 0:
                        chk_btn1.configure(state=DISABLED)
                        emp_prod_profit.append([str(entry_2.get())])
                        emp_profit = 0
                        tot_prod_profit = 0
                        for index1 in tree_view1.get_children():
                            for index2 in tree_view2.get_children():
                                prod_real = tree_view1.item(index1)['values']
                                prod_img = tree_view2.item(index2)['values']
                                if prod_real[0] == prod_img[0]:
                                    emp_profit = emp_profit + float(prod_real[2]) * float(
                                        prod_img[1][:prod_img[1].index('%')]) / 100
                                    break
                        print("EMP_PROFIT: ", emp_profit)
                        print("EMP_PROD_PROFIT: ", emp_prod_profit)
                        for item in emp_prod_profit:
                            print(item, " == ", entry_2.get(), "\nTypes: ", type(item), " ?= ", type(entry_2.get()))
                            if item[0] == str(entry_2.get()):
                                emp_prod_profit[emp_prod_profit.index(item)].append(emp_profit)
                        print("EMP_PROD_PROFIT: ", emp_prod_profit)
                        for index1 in tree_view1.get_children():
                            tot_prod_profit += float(tree_view1.item(index1)['values'][2])
                        for item in emp_prod_profit:
                            if item[0] == str(entry_2.get()):
                                emp_prod_profit[emp_prod_profit.index(item)].append(round_to_3(tot_prod_profit))
                                break
                        print("EMP_PROD_PROFIT: ", emp_prod_profit)
                        final_prod_profit = 0
                        for i in range(len(emp_prod_profit)):
                            final_prod_profit += emp_prod_profit[i][2]
                            i += 1
                        lbl_3.configure(text=final_prod_profit)
                        if entry_1.get() != '':
                            col_excluding_prof = float(entry_1.get()) - float(lbl_3.cget('text'))
                            lbl_5.configure(text=col_excluding_prof)
                        print("EMP_PROD_PROFIT: ", emp_prod_profit)
            else:
                fill_data_msg(1, "Data Incomplete", "Please fill all the required details.")

    def rem_select_rec_tree(a):
        global products, emp_prod_profit
        if a == 1:
            records = tree_view1.selection()
            for record in records:
                tree_view1.delete(record)
        if a == 2:
            records = tree_view2.selection()
            for record in records:
                deletable = tree_view2.item(record)["values"][0]
                tree_view2.delete(record)
                for child_tr1 in tree_view1.get_children():
                    if tree_view1.item(child_tr1)['values'][0] == deletable:
                        tree_view1.delete(child_tr1)
                products.remove(deletable)
                cmb_box1.set('')
                cmb_box1.config(values=products)
        if a == 3:
            if tree_view3.get_children() != ():
                lbl_20.configure(fg="yellow")
            records = tree_view3.selection()
            no_prods = True
            for record in records:
                for item in emp_prod_profit:
                    if item[0] == tree_view3.item(record)['values'][0]:
                        emp_prod_profit.remove(item)
                tree_view3.delete(record)
                final_prod_profit = 0
                for i in range(len(emp_prod_profit)):
                    final_prod_profit += emp_prod_profit[i][2]
                    i += 1
                lbl_3.configure(text=final_prod_profit)
            if entry_1.get() != '':
                col_excluding_prof = float(entry_1.get()) - float(lbl_3.cget('text'))
                lbl_5.configure(text=col_excluding_prof)
            print(emp_prod_profit)
            for index in tree_view3.get_children():
                if tree_view3.item(index)['values'][3] != 0:
                    no_prods = False
                    break
            if no_prods:
                chk_btn1.configure(state=NORMAL)
            if tree_view3.get_children() == ():
                lbl_16.configure(text='-')

    def rem_all_rec_tree(a):
        global products, emp_prod_profit
        if a == 1:
            for record in tree_view1.get_children():
                tree_view1.delete(record)
        if a == 2:
            for record in tree_view2.get_children():
                products.clear()
                cmb_box1.config(values=products)
                cmb_box1.set('')
                tree_view2.delete(record)
                rem_all_rec_tree(1)
        if a == 3:
            if tree_view3.get_children() != ():
                lbl_20.configure(fg="yellow")
            for record in tree_view3.get_children():
                tree_view3.delete(record)
            emp_prod_profit.clear()
            chk_btn1.configure(state=NORMAL)
            lbl_3.configure(text='-')
            if entry_1.get() == '':
                lbl_5.configure(text='-')
            else:
                lbl_5.configure(text=entry_1.get())
            lbl_16.configure(text='-')

    def round_to_3(data):
        return round(data, 3)

    def validate_if_float(new_value):
        if new_value == "":
            return True
        try:
            float(new_value)
            if new_value.count(".") > 1:
                return False
            return True
        except ValueError:
            return False

    def validate_percent(new_value):
        if new_value == "" or new_value == placeholder_ent4a or new_value == placeholder_ent4b:
            return True
        try:
            float_val = float(new_value)
            if new_value.count(".") > 1:
                return False
            elif float_val < 1 or float_val > 100:
                fill_data_msg(2, "Out of Limit", "Percentage can be between 1 to 100 only.")
                return False
            return True
        except ValueError:
            return False

    def validate_if_num(new_value):
        if new_value.isdigit() or new_value == "":
            return True
        else:
            return False

    def on_child_min(event):
        if event.widget == child:
            root.deiconify()

    global window_counter1
    global products
    if window_counter1 < 1:
        window_counter1 += 1
        root.iconify()
        child = Toplevel()
        child.title("Create A Record")
        child.resizable(False, False)
        child.geometry(f"{child.winfo_screenwidth()}x{child.winfo_screenheight()}+-7+0")
        child.configure(bg="#313232")
        child.protocol("WM_DELETE_WINDOW", close_action)
        child.bind("<Unmap>", on_child_min)
        child.iconbitmap("Required Images/app_icon.ico")

        validate1 = child.register(validate_if_float)
        validate2 = child.register(validate_percent)
        validate3 = child.register(validate_if_num)

        frame1 = LabelFrame(child, fg='white', bg="#203354", text="Employee Details")
        frame1.place(x=40, y=110, width=653, height=280)
        frame2 = LabelFrame(child, fg='white', bg="#203354", text="Product Details")
        frame2.place(x=740, y=40, width=606, height=320)
        tree_frame1 = Frame(child)
        tree_frame1.place(x=340, y=120, width=180, height=209)
        tree_frame2 = Frame(child)
        tree_frame2.place(x=1000, y=50, width=330, height=298)
        tree_frame3 = Frame(child, bg="#b4b4b4")
        tree_frame3.place(x=40, y=388, width=927, height=330)

        scr1a = Scrollbar(tree_frame1, width=15)
        scr1a.pack(side=RIGHT, fill=Y)
        scr1b = Scrollbar(tree_frame1, orient=HORIZONTAL, width=15)
        scr1b.pack(side=BOTTOM, fill=X)
        scr2a = Scrollbar(tree_frame2, width=15)
        scr2a.pack(side=RIGHT, fill=Y)
        scr2b = Scrollbar(tree_frame2, orient=HORIZONTAL, width=15)
        scr2b.pack(side=BOTTOM, fill=X)
        scr3a = Scrollbar(tree_frame3, width=15)
        scr3a.pack(side=RIGHT, fill=Y)

        tree_view1 = ttk.Treeview(tree_frame1, yscrollcommand=scr1a.set, xscrollcommand=scr1b.set)
        tree_view1["columns"] = ("Product", "Quantity", "Total Profit")
        tree_view1.column("#0", width=0, stretch=NO)
        tree_view1.column("Product", anchor=W, width=80)
        tree_view1.column("Quantity", anchor=W, width=80)
        tree_view1.column("Total Profit", anchor=W, width=80)
        tree_view1.heading("Product", text="Product", anchor=W)
        tree_view1.heading("Quantity", text="Quantity", anchor=W)
        tree_view1.heading("Total Profit", text="Total Profit", anchor=W)
        tree_view1.pack(side=LEFT, expand=True, fill=BOTH)
        tree_view1.tag_configure('bluerow', background='lightblue')
        scr1a.config(command=tree_view1.yview)
        scr1b.config(command=tree_view1.xview)

        tree_view2 = ttk.Treeview(tree_frame2, yscrollcommand=scr2a.set, xscrollcommand=scr2b.set)
        tree_view2["columns"] = ("Product Name", "Employee Margin", "Profit")
        tree_view2.column("#0", width=0, stretch=NO)
        tree_view2.column("Product Name", anchor=W, width=140)
        tree_view2.column("Employee Margin", anchor=W, width=160)
        tree_view2.column("Profit", anchor=W, width=120)
        tree_view2.heading("Product Name", text="Product Name", anchor=W)
        tree_view2.heading("Employee Margin", text="Employee Margin", anchor=W)
        tree_view2.heading("Profit", text="Profit", anchor=W)
        tree_view2.pack(side=LEFT, expand=True, fill=BOTH)
        tree_view2.tag_configure('bluerow', background='lightblue')
        scr2a.config(command=tree_view2.yview)
        scr2b.config(command=tree_view2.xview)

        tree_view3 = ttk.Treeview(tree_frame3, yscrollcommand=scr3a.set)
        tree_view3["columns"] = ("ID", "Name", "Salary", "Products Sold",
                                 "Estimated Income")
        tree_view3.column("#0", width=0, stretch=NO)
        tree_view3.column("ID", anchor=W, width=80)
        tree_view3.column("Name", anchor=W, width=80)
        tree_view3.column("Salary", anchor=W, width=80)
        tree_view3.column("Products Sold", anchor=W, width=80)
        tree_view3.column("Estimated Income", anchor=W, width=80)
        tree_view3.heading("ID", text="ID", anchor=W)
        tree_view3.heading("Name", text="Name", anchor=W)
        tree_view3.heading("Salary", text="Salary", anchor=W)
        tree_view3.heading("Products Sold", text="Products Sold", anchor=W)
        tree_view3.heading("Estimated Income", text="Estimated Income", anchor=W)
        tree_view3.pack(side=LEFT, expand=True, fill=BOTH)
        tree_view3.tag_configure('bluerow', background='lightblue')
        scr3a.config(command=tree_view3.yview)

        style = ttk.Style()
        style.theme_use("default")
        style.configure("Treeview",
                        background="#white",
                        foreground="black",
                        rowheight=25,
                        fieldbackground="white")
        style.map('Treeview', background=[('selected', 'blue')])

        btn_1 = Button(child, bg="#800000", fg="white", text="Save File", font=("Times New Roman", 20, "bold"), command=save_file)
        btn_1.place(x=990, y=655, width=234, height=65)
        btn_2 = Button(child, bg="#203354", fg="white", text="Estimate Incomes", font=("Times New Roman", 22, "bold"), command=estimate_inc)
        btn_2.place(x=990, y=390, width=234, height=52)
        btn_3 = Button(child, bg="#03A9F4", text="Add Record", command=lambda: add_rec_tree(3))
        btn_3.place(x=50, y=350, width=119, height=30)
        btn_4 = Button(child, bg="#03A9F4", text="Update Selected", command=lambda: upd_rec_tree(3))
        btn_4.place(x=178, y=350, width=119, height=30)
        btn_5 = Button(child, bg="#03A9F4", text="Remove Selected", command=lambda: rem_select_rec_tree(3))
        btn_5.place(x=306, y=350, width=119, height=30)
        btn_6 = Button(child, bg="#03A9F4", text="Remove All", command=lambda: rem_all_rec_tree(3))
        btn_6.place(x=435, y=350, width=119, height=30)
        btn_7 = Button(child, bg="#03A9F4", text="Add Record", state=DISABLED, command=lambda: add_rec_tree(1))
        btn_7.place(x=540, y=220, width=120, height=20)
        btn_8 = Button(child, bg="#03A9F4", text="Update Selected", state=DISABLED, command=lambda: upd_rec_tree(1))
        btn_8.place(x=540, y=250, width=120, height=20)
        btn_9 = Button(child, bg="#03A9F4", text="Remove Selected", state=DISABLED, command=lambda: rem_select_rec_tree(1))
        btn_9.place(x=540, y=280, width=120, height=20)
        btn_10 = Button(child, bg="#03A9F4", text="Remove All", state=DISABLED, command=lambda: rem_all_rec_tree(1))
        btn_10.place(x=540, y=310, width=120, height=20)
        btn_11 = Button(child, bg="#03A9F4", text="Add Product", state=DISABLED, command=lambda: add_rec_tree(2))
        btn_11.place(x=760, y=200, width=211, height=30)
        btn_12 = Button(child, bg="#03A9F4", text="Update Selected Product", state=DISABLED, command=lambda: upd_rec_tree(2))
        btn_12.place(x=760, y=240, width=211, height=30)
        btn_13 = Button(child, bg="#03A9F4", text="Remove Selected Products", state=DISABLED, command=lambda: rem_select_rec_tree(2))
        btn_13.place(x=760, y=280, width=211, height=30)
        btn_14 = Button(child, bg="#03A9F4", text="Remove All", state=DISABLED, command=lambda: rem_all_rec_tree(2))
        btn_14.place(x=760, y=320, width=211, height=30)
        btn_15 = Button(child, bg="#494949", image=chat_img, font=f, command=open_chatbot)
        btn_15.place(x=1244, y=655, width=75, height=60)
        btn_16 = Button(child, bg="#800000", fg="white", text="Set/Reset", font=("Times New Roman", 15, 'bold'), command=set_reset)
        btn_16.place(x=350, y=1, width=155, height=30)

        v1 = IntVar()
        chk_btn1 = Checkbutton(child, fg="white", bg="#202030", text="Include Products", variable=v1, offvalue=0, onvalue=1, command=enable_prod)
        chk_btn1["font"] = ("Calibre", 15, "bold")
        chk_btn1["selectcolor"] = "black"
        chk_btn1.place(x=475, y=40, width=218, height=38)

        v2 = IntVar()
        rad_btn1 = Radiobutton(child, selectcolor='black', fg='white', text="Fixed Amount", bg="#203354", variable=v2, value=1, command=choose)
        rad_btn1.place(x=130, y=220, width=99, height=30)
        rad_btn2 = Radiobutton(child, selectcolor='black', fg='white', text="Percentage Based", bg="#203354", variable=v2, value=2, command=choose)
        rad_btn2.place(x=128, y=250, width=122, height=30)

        lbl_1 = Label(child, fg="white", bg="#313232", text="Enter total collection:", font=('Caliber', 10, 'bold'))
        lbl_1.place(x=35, y=0, width=150, height=30)
        lbl_2 = Label(child, fg="white", bg="#313232", text="Total Profit from Product Sale:", font=('Caliber', 10, 'bold'), state=DISABLED)
        lbl_2.place(x=30, y=40, width=216, height=30)
        lbl_3 = Label(child, fg="#FFF157", bg="#313232", text="-", font=('Caliber', 15, 'bold'), state=DISABLED)
        lbl_3.place(x=242, y=40, width=102, height=30)
        lbl_4 = Label(child, fg="white", bg="#313232", text="Collection excluding Product Sale:", font=('Caliber', 10, 'bold'))
        lbl_4.place(x=42, y=70, width=220, height=30)
        lbl_5 = Label(child, fg="#FFF157", bg="#313232", text="-", font=('Caliber', 15, 'bold'))
        lbl_5.place(x=272, y=70, width=111, height=31)
        lbl_6 = Label(child, fg='white', bg="#203354", text="ID:", font=f)
        lbl_6.place(x=41, y=130, width=61, height=37)
        lbl_7 = Label(child, fg='white', bg="#203354", text="Name:", font=f)
        lbl_7.place(x=41, y=170, width=82, height=32)
        lbl_8 = Label(child, fg='white', bg="#203354", text="Salary:", font=f)
        lbl_8.place(x=42, y=208, width=82, height=39)
        lbl_9 = Label(child, fg='white', bg="#203354", text="Product:", font=f, state=DISABLED)
        lbl_9.place(x=520, y=130, width=70, height=25)
        lbl_10 = Label(child, fg='white', bg="#203354", text="Quantity:", font=f, state=DISABLED)
        lbl_10.place(x=520, y=170, width=70, height=25)
        lbl_11 = Label(child, fg='white', bg="#203354", text="Product Name:", state=DISABLED, font=f)
        lbl_11.place(x=740, y=60, width=108, height=37)
        lbl_12 = Label(child, fg='white', bg="#203354", text="Employee Margin:", state=DISABLED, font=f)
        lbl_12.place(x=740, y=90, width=126, height=38)
        lbl_13 = Label(child, fg='white', bg="#203354", text="Cost Price:", font=f, state=DISABLED)
        lbl_13.place(x=740, y=120, width=83, height=36)
        lbl_14 = Label(child, fg='white', bg="#203354", text="Selling Price:", font=f, state=DISABLED)
        lbl_14.place(x=740, y=150, width=96, height=39)
        lbl_15 = Label(child, fg="white", bg="#313232", text="Amount left after Distribution:", font=("Roboto", 13, "bold"))
        lbl_15.place(x=990, y=480, width=262, height=30)
        lbl_16 = Label(child, fg="#32CD32", bg="#313232", text="-", font=("Times New Roman", 20, "bold"))
        lbl_16.place(x=990, y=510, width=239, height=38)
        lbl_17 = Label(child, bg="white")
        lbl_17.place(x=40, y=340, width=652, height=1)
        lbl_17 = Label(child, bg="white")
        lbl_17.place(x=40, y=341, width=652, height=1)
        lbl_18 = Label(child, fg="white", bg="#313232", text="Enter File Name:",
                       font=("Roboto", 13, "bold"))
        lbl_18.place(x=990, y=560, width=232, height=30)
        lbl_19 = Label(child, fg="white", bg="#313232", text=".xlsx",
                       font=("Caliber", 13, "bold"))
        lbl_19.place(x=1227, y=600, width=50, height=30)
        lbl_20 = Label(child, fg="#313232", bg="#313232", justify=LEFT, text="<--\nValues\nUpdated !", font=("Roboto", 13, 'bold'))
        lbl_20.place(x=1230, y=390, width=115, height=55)

        entry_1 = Entry(child, justify=CENTER, validate="key", validatecommand=(validate1, "%P"))
        entry_1.place(x=183, y=1, width=155, height=30)
        entry_2 = Entry(child, justify=CENTER)
        entry_2.place(x=120, y=140, width=140, height=20)
        entry_3 = Entry(child, justify=CENTER)
        entry_3.place(x=120, y=180, width=140, height=20)
        entry_4 = Entry(child, fg="#b4b4b4", justify=CENTER, state=DISABLED, validate="key")
        placeholder_ent4a = "Enter Amount"
        placeholder_ent4b = "Enter % Collection"
        entry_4.bind("<FocusIn>", when_focused1)
        entry_4.place(x=120, y=290, width=135, height=30)
        entry_5 = Entry(child, justify=CENTER, state=DISABLED, validate="key", validatecommand=(validate3, "%P"))
        entry_5.place(x=590, y=170, width=79, height=30)
        entry_6 = Entry(child, justify=CENTER, state=DISABLED)
        entry_6.place(x=860, y=67, width=125, height=20)
        entry_7 = Entry(child, justify=CENTER, state=DISABLED, validate="key", validatecommand=(validate2, "%P"))
        entry_7.place(x=860, y=100, width=125, height=20)
        entry_8 = Entry(child, justify=CENTER, state=DISABLED, validate="key", validatecommand=(validate1, "%P"))
        entry_8.place(x=860, y=130, width=125, height=20)
        entry_9 = Entry(child, justify=CENTER, state=DISABLED, validate="key", validatecommand=(validate1, "%P"))
        entry_9.place(x=860, y=160, width=125, height=20)
        entry_10 = Entry(child, justify=CENTER)
        entry_10.place(x=990, y=600, width=232, height=30)

        cmb_box1 = ttk.Combobox(child, values=[], state=DISABLED)
        cmb_box1.place(x=590, y=130, width=80, height=25)
        cmb_box1.bind("<Button-1>", on_combo_select)

        child.mainloop()
    else:
        messagebox.showinfo("Limit reached!", "Can Open Only One Window")


f1 = ("Times New Roman", 20, "bold")
f = ("Calibre", 10)

window_counter1 = 0
count_tree1 = 0
count_tree2 = 0
count_tree3 = 0
products = []
emp_prod_profit = []
root = Tk()
root.title("Income Estimator")
root.geometry("540x440+0+0")
root.resizable(False, False)
root.iconbitmap("Required Images/app_icon.ico")
root.protocol("WM_DELETE_WINDOW", close_win)
img = ImageTk.PhotoImage(Image.open("Required Images/MainWindowImage.jpg"))
chat_img = ImageTk.PhotoImage(Image.open("Required Images/chat_image.png"))
img_lbl = Label(image=img)
img_lbl.pack()
title_lbl = Label(root, text="Income Estimator", bg="#800000", font=("Impact", 40), fg="#FFD700", justify="center")
title_lbl.place(x=36, y=20, width=465, height=85)
crt_new = Button(root, fg="white", bg="#313232", text="Create New File", font=f1, justify="center", command=open_create)
crt_new.place(x=116, y=230, width=305, height=30)
src_old = Button(root, fg="white", bg="#313232", text="View Existing File", font=f1, justify="center", command=browse_folder)
src_old.place(x=116, y=280, width=306, height=30)
chat_bot = Button(root, bg="#494949", image=chat_img, font=f, command=open_chatbot)
chat_bot.place(x=460, y=380, width=69, height=51)


root.mainloop()
